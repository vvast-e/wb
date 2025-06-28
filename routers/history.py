import csv
from datetime import datetime

import pandas as pd
from fastapi import Depends, APIRouter, HTTPException, status, Query
from fastapi.responses import StreamingResponse
import io

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional

from database import get_db
from dependencies import get_current_user_with_wb_key, get_wb_api_key
from models import User
from crud.history import get_tasks, revert_history_change_by_id, revert_history_media

router = APIRouter(tags=["History"], prefix="/api/history")


@router.get("/", response_model=dict)
async def read_history(
        page: int = Query(1, ge=1),
        per_page: int = Query(20, ge=1, le=100),
        order_by: str = Query("created_at",
                              regex="^(created_at|scheduled_at|vendor_code|brand|action|status|user_email)$"),
        order_dir: str = Query("desc", regex="^(asc|desc)$"),
        email: Optional[str] = Query(None),
        brand: str = Query(..., description="Название бренда"),
        vendor_code: Optional[str] = Query(None),
        date_from: Optional[datetime] = Query(None),
        date_to: Optional[datetime] = Query(None),
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user_with_wb_key),

):
    offset = (page - 1) * per_page
    histories, total = await get_tasks(
        db,
        offset=offset,
        limit=per_page,
        order_by=order_by,
        order_dir=order_dir,
        email=email,
        brand=brand,
        vendor_code=vendor_code,
        date_from=date_from,
        date_to=date_to
    )

    if not histories:
        raise HTTPException(status_code=404, detail="История не найдена")

    return {
        "items": histories,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page
    }


@router.get("/export")
async def export_history(
        format: str = Query("csv", pattern="^(csv|excel)$"),
        email: Optional[str] = None,
        brand: Optional[str] = None,
        vendor_code: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user_with_wb_key)
):
    histories, _ = await get_tasks(
        db,
        offset=0,
        limit=None,
        email=email,
        brand=brand,
        vendor_code=vendor_code,
        date_from=date_from,
        date_to=date_to
    )

    if not histories:
        raise HTTPException(status_code=404, detail="Нет данных для экспорта")

    # Подготовка данных
    data = []
    for item in histories:
        readable_changes = ", ".join([f"{k}: {v}" for k, v in (item.changes or {}).items()])
        data.append({
            "Дата создания": item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "Дата выполнения": item.scheduled_at.strftime('%Y-%m-%d %H:%M:%S') if item.scheduled_at else "",
            "Артикул": item.vendor_code,
            "Бренд": item.brand,
            "Действие": item.action,
            "Изменения": readable_changes,
            "Пользователь": item.user.email if item.user else "",
            "Статус": item.status
        })

    if format == "csv":
        df = pd.DataFrame(data)
        buffer = io.BytesIO()
        df.to_csv(buffer, index=False, encoding="utf-8-sig", sep=';')
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=history_export.csv"}
        )

    elif format == "excel":
        df = pd.DataFrame(data)
        output = io.BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        # Открываем workbook и настраиваем ширину столбцов
        workbook = load_workbook(output)
        sheet = workbook.active
        for col_idx, column_cells in enumerate(sheet.iter_cols(1, sheet.max_column), 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

        final_output = io.BytesIO()
        workbook.save(final_output)
        final_output.seek(0)

        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=history_export.xlsx"}
        )


@router.post("/{history_id}/revert")
async def revert_history_change(
        history_id: int,
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user_with_wb_key),
):
    await revert_history_change_by_id(db, history_id)
    return {"detail": "Откат изменений запланирован"}


@router.post("/{history_id}/revert-media")
async def revert_media_change(
        history_id: int,
        db: AsyncSession = Depends(get_db),
        wb_api_key: str = Depends(get_wb_api_key),
        current_user: User = Depends(get_current_user_with_wb_key),
):
    await revert_history_media(db, history_id, wb_api_key)
    return {"detail": "Откат изменений медиа запланирован"}
